import io
import json
import os
import sqlite3
import sys
import uuid
from datetime import datetime, date

from flask import Flask, render_template, request, redirect, url_for, flash, g, send_file, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

DB_PATH = None
IMPORT_TMP_DIR = None

# ---------------- Konfigurasi Path untuk PyInstaller Exe ----------------
if getattr(sys, "frozen", False):
    # Jika berjalan sebagai .exe hasil build PyInstaller
    BASE_DIR = sys._MEIPASS
    # Letakkan database di folder yang sama dengan file .exe utama
    DB_PATH = os.path.join(os.path.dirname(sys.executable), "sidap.db")
else:
    # Jika berjalan normal via python app.py
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DB_PATH = os.path.join(BASE_DIR, "sidap.db")

IMPORT_TMP_DIR = os.path.join(BASE_DIR, "import_tmp")

# Beritahu Flask lokasi folder templates dan static yang berada di dalam folder ekstraksi .exe
app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static"),
)
app.secret_key = "sidap-dev-secret"


def is_valid_iso_date(s):
    """Cek strict format YYYY-MM-DD DAN kewajaran tahunnya. Jangan percaya
    <input type=date> di browser bakal selalu ngirim format yang masuk akal
    -- interaksi keyboard yang aneh di date picker bisa hasilin tahun kayak
    '4423' yang formatnya sah tapi jelas typo."""
    try:
        parsed = datetime.strptime(s, "%Y-%m-%d")
    except (ValueError, TypeError):
        return False
    tahun_sekarang = datetime.now().year
    return 1990 <= parsed.year <= tahun_sekarang + 2

# ---------------- Koneksi DB ----------------
def get_db():
    if "db" not in g:
        if not db_has_schema():
            init_db()
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        

        normalisasi_database_kecamatan(g.db)
        
    return g.db

def normalisasi_database_kecamatan(db):
    """Membersihkan variasi nama kecamatan yang sudah terlanjur tersimpan di DB."""
    try:
        # 1. Bersihkan tabel permohonan
        rows_perm = db.execute("SELECT DISTINCT kecamatan FROM permohonan WHERE kecamatan IS NOT NULL").fetchall()
        for r in rows_perm:
            kec_lama = r["kecamatan"]
            kec_baru = bersihkan_kecamatan(kec_lama)
            if kec_lama != kec_baru:
                db.execute("UPDATE permohonan SET kecamatan=? WHERE kecamatan=?", (kec_baru, kec_lama))

        # 2. Bersihkan tabel pelanggan
        rows_pel = db.execute("SELECT DISTINCT kecamatan FROM pelanggan WHERE kecamatan IS NOT NULL").fetchall()
        for r in rows_pel:
            kec_lama = r["kecamatan"]
            kec_baru = bersihkan_kecamatan(kec_lama)
            if kec_lama != kec_baru:
                db.execute("UPDATE pelanggan SET kecamatan=? WHERE kecamatan=?", (kec_baru, kec_lama))

        db.commit()
    except Exception as e:
        print("Peringatan normalisasi kecamatan:", e)

@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    schema_path = os.path.join(BASE_DIR, "schema.sql")
    if not os.path.exists(schema_path):
        raise FileNotFoundError(f"schema.sql not found in {BASE_DIR}. Please restore schema.sql before running SIDAP.")

    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    with open(schema_path, "r", encoding="utf-8") as f:
        conn.executescript(f.read())
    conn.commit()
    conn.close()


def db_has_schema():
    if not os.path.exists(DB_PATH):
        return False
    try:
        conn = sqlite3.connect(DB_PATH)
        required_tables = ["pelanggan", "instalasi", "permohonan"]
        existing = {
            row[0] for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }
        conn.close()
        return all(name in existing for name in required_tables)
    except sqlite3.DatabaseError:
        return False


def ensure_data_paths():
    os.makedirs(IMPORT_TMP_DIR, exist_ok=True)
    if not db_has_schema():
        init_db()


def daftar_kecamatan(db):
    rows = db.execute(
        "SELECT DISTINCT kecamatan FROM pelanggan ORDER BY kecamatan"
    ).fetchall()
    return [r["kecamatan"] for r in rows]


def daftar_pelanggan_pilihan(db):
    rows = db.execute("SELECT id, nama, kecamatan FROM pelanggan ORDER BY nama").fetchall()
    return [{"id": r["id"], "label": f"{r['nama']} \u2014 {r['kecamatan']}"} for r in rows]


# ---------------- Import Excel: template + parsing ----------------
HEADER_MAP = {
    "no": "no",
    "nama pelanggan": "nama",
    "alamat pelanggan": "alamat",
    "kelurahan": "kelurahan",
    "kecamatan": "kecamatan",
    "nomor instalasi": "nomor_instalasi",
    "tanggal pasang": "tanggal_pasang",
    "diameter pipa": "diameter_pipa",
    "tekanan air": "tekanan_air",
    "status (sib/bk)": "status",
    "petugas": "petugas",
    "keterangan": "keterangan",
}
WAJIB_DIISI = ["nama", "kecamatan", "nomor_instalasi", "tanggal_pasang", "status"]
VALID_STATUSES = ("SIB", "BK")

PERMOHONAN_HEADER_MAP = {
    "no": "no",
    "nama pelanggan": "nama_pelanggan",
    "alamat": "lokasi",
    "kelurahan": "kelurahan",
    "kecamatan": "kecamatan",
    "no. spk": "no_spk",
    "no spk": "no_spk",
    "ke perencana": "tanggal_permohonan",
    "tanggal permohonan": "tanggal_permohonan",
    "survey": "tanggal_survey",
    "tanggal survey": "tanggal_survey",
    "kembali ke hublang": "tanggal_kembali_hublang",
    "petugas survey": "petugas_survey",
    "petugas": "petugas_survey",
    "keterangan": "keterangan",
}
PERMOHONAN_REQUIRED = ["nama_pelanggan", "kecamatan", "jenis", "tanggal_permohonan"]
PERMOHONAN_PIPA_TYPES = ("P.Dinas", "P.Distribusi")
PERMOHONAN_DITINDAKLANJUTI = {
    "": None,
    "belum": None,
    "ya": 1,
    "tidak": 0,
    "0": 0,
    "1": 1,
}
KECAMATAN_MAP = {
    # Variasi Panombean Panei
    "p. pane": "Panombean Panei",
    "p.pane": "Panombean Panei",
    "p. panei": "Panombean Panei",
    "p.panei": "Panombean Panei",
    "pane": "Panombean Panei",
    "panei": "Panombean Panei",
    "panombean pane": "Panombean Panei",
    "panombean panei": "Panombean Panei",
    "s. panei": "Panombean Panei",
    "s.panei": "Panombean Panei",
    "s. pane": "Panombean Panei",
    "s.pane": "Panombean Panei",
    "S. P.Panei" : "Panombean Panei",

    # Variasi Siantar Barat
    "s. barat": "Siantar Barat",
    "s.barat": "Siantar Barat",
    "siantar barat": "Siantar Barat",

    # Variasi Siantar Marihat
    "s. marihat": "Siantar Marihat",
    "s.marihat": "Siantar Marihat",
    "siantar marihat": "Siantar Marihat",

    # Variasi Siantar Marimbun
    "s. marimbun": "Siantar Marimbun",
    "s.marimbun": "Siantar Marimbun",
    "siantar marimbun": "Siantar Marimbun",
    "S.Simarimbun": "Siantar Marimbun",

    # Variasi Siantar Martoba
    "s. martoba": "Siantar Martoba",
    "s.martoba": "Siantar Martoba",
    "siantar martoba": "Siantar Martoba",

    # Variasi Siantar Sitalasari
    "s. sitalasari": "Siantar Sitalasari",
    "s.sitalasari": "Siantar Sitalasari",
    "siantar sitalasari": "Siantar Sitalasari",

    # Variasi Siantar Selatan
    "s. selatan": "Siantar Selatan",
    "s.selatan": "Siantar Selatan",
    "siantar selatan": "Siantar Selatan",

    # Variasi Siantar Timur
    "s. timur": "Siantar Timur",
    "s.timur": "Siantar Timur",
    "siantar timur": "Siantar Timur",

    # Variasi Siantar Utara
    "s. utara": "Siantar Utara",
    "s.utara": "Siantar Utara",
    "siantar utara": "Siantar Utara",

    # Variasi Siantar Simalungun / Simalungun
    "s. malungun": "Siantar Simalungun",
    "s.malungun": "Siantar Simalungun",
    "s. simalungun": "Siantar Simalungun",
    "s.simalungun": "Siantar Simalungun",
}

def bersihkan_kecamatan(raw_name):
    """Mengubah variasi singkatan/typo kecamatan menjadi nama resmi standar."""
    if not raw_name:
        return ""
    cleaned = str(raw_name).strip().lower()
    return KECAMATAN_MAP.get(cleaned, raw_name.strip().title())


def generate_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pelanggan"

    headers = [
        "No", "Nama Pelanggan", "Alamat Pelanggan", "Kelurahan", "Kecamatan",
        "Nomor Instalasi", "Tanggal Pasang", "Diameter Pipa", "Tekanan Air",
        "Status (SIB/BK)", "Petugas", "Keterangan",
    ]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="14203A", end_color="14203A")
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D7DAE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 30

    widths = [6, 22, 26, 16, 16, 16, 14, 12, 12, 16, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    example = [1, "Contoh Nama Pelanggan", "Jl. Contoh No.1", "Contoh Kelurahan",
               "Contoh Kecamatan", "83.001", "2026-01-15", "\u00d82", 0.4, "SIB", "", ""]
    for col, val in enumerate(example, start=1):
        c = ws.cell(row=2, column=col, value=val)
        c.font = body_font
        c.border = border

    ws.cell(row=2, column=6).number_format = "@"
    for r in range(2, 202):
        ws.cell(row=r, column=6).number_format = "@"
        ws.cell(row=r, column=7).number_format = "DD-MM-YYYY"
        for col in range(1, 13):
            ws.cell(row=r, column=col).border = border
            ws.cell(row=r, column=col).font = body_font

    dv = DataValidation(type="list", formula1='"SIB,BK"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("J2:J201")
    ws.freeze_panes = "A2"

    notes = wb.create_sheet("Catatan Pengisian")
    rows = [
        ("Kolom", "Catatan"),
        ("Nomor Instalasi", "Ketik apa adanya termasuk titik (contoh: 83.236). Jangan biarkan Excel mengubahnya jadi angka."),
        ("Nama Pelanggan", "Kalau satu pelanggan punya lebih dari satu instalasi, tulis nama persis sama di tiap baris -- itu bukan duplikat."),
        ("Status (SIB/BK)", "Pilih dari dropdown. SIB = Sambungan Instalasi Baru, BK = Buka Kembali."),
        ("Petugas", "Boleh dikosongkan."),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        notes.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
        cb = notes.cell(row=r, column=2, value=b)
        cb.font = Font(bold=(r == 1))
        cb.alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 20
    notes.column_dimensions["B"].width = 75

    return wb


def parse_tanggal(val):
    if val is None or val == "":
        return None, None
    if isinstance(val, (datetime, date)):
        parsed = val if isinstance(val, datetime) else datetime.combine(val, datetime.min.time())
        if not (1990 <= parsed.year <= datetime.now().year + 2):
            return None, f'Tahun tanggal "{parsed.year}" tidak wajar'
        return parsed.strftime("%Y-%m-%d"), None
    raw = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y"):
        try:
            parsed = datetime.strptime(raw, fmt)
            if not (1990 <= parsed.year <= datetime.now().year + 2):
                return None, f'Tahun tanggal "{parsed.year}" tidak wajar'
            return parsed.strftime("%Y-%m-%d"), None
        except ValueError:
            continue
    return None, f'Format tanggal "{raw}" tidak dikenali, pakai format tanggal (bukan teks bebas)'


def validate_permohonan_data(data):
    errors = []

    if not data.get("nama_pelanggan", "").strip():
        errors.append("Nama pelanggan kosong")
    if not data.get("kecamatan", "").strip():
        errors.append("Kecamatan kosong")
    if data.get("jenis") not in VALID_STATUSES:
        errors.append("Jenis permohonan tidak valid")
    if not data.get("tanggal_permohonan", "").strip():
        errors.append("Tanggal permohonan kosong")
    elif not is_valid_iso_date(data.get("tanggal_permohonan")):
        errors.append("Tanggal permohonan tidak valid")

    tanggal_survey = data.get("tanggal_survey", "").strip()
    if tanggal_survey and not is_valid_iso_date(tanggal_survey):
        errors.append("Tanggal survey tidak valid")

    tanggal_dikirim = data.get("tanggal_dikirim_hublang", "").strip()
    if tanggal_dikirim and not is_valid_iso_date(tanggal_dikirim):
        errors.append("Tanggal dikirim ke Hublang tidak valid")

    tanggal_kembali = data.get("tanggal_kembali_hublang", "").strip()
    if tanggal_kembali and not is_valid_iso_date(tanggal_kembali):
        errors.append("Tanggal kembali ke Hublang tidak valid")

    return errors


def parse_upload(fileobj, db):
    wb = load_workbook(fileobj, data_only=True)
    if not wb.worksheets:
        raise ValueError("File Excel ini kosong, tidak ada sheet sama sekali.")
    # Pakai sheet PERTAMA di file, apapun namanya -- jangan maksa nama sheet
    # persis "Data Pelanggan". Kalau orang rename sheet, save-as, atau
    # duplikat tab pas ngedit manual, import tetap harus jalan selama
    # struktur kolomnya benar.
    ws = wb.worksheets[0]

    header_row = [c.value for c in ws[1]]
    col_index = {}
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        key = HEADER_MAP.get(str(h).strip().lower())
        if key:
            col_index[key] = idx

    missing = [k for k in ["nama", "kecamatan", "nomor_instalasi", "tanggal_pasang", "status"] if k not in col_index]
    if missing:
        raise ValueError(f"Kolom wajib tidak ditemukan di header: {', '.join(missing)}")

    existing_nomor = {
        r["nomor_instalasi"] for r in db.execute("SELECT nomor_instalasi FROM instalasi").fetchall()
    }
    existing_pelanggan = {}
    for r in db.execute("SELECT id, nama, kecamatan FROM pelanggan").fetchall():
        existing_pelanggan[(r["nama"].strip().lower(), r["kecamatan"].strip().lower())] = r["id"]

    rows = []
    batch_pelanggan = {}
    next_batch_no = 1

    for row_cells in ws.iter_rows(min_row=2, values_only=False):
        values = {}
        for key, idx in col_index.items():
            values[key] = row_cells[idx].value if idx < len(row_cells) else None

        nama = str(values.get("nama") or "").strip()
        kecamatan_raw = str(values.get("kecamatan") or "").strip()
        kecamatan = bersihkan_kecamatan(kecamatan_raw)
        nomor_instalasi = str(values.get("nomor_instalasi") or "").strip()
        tanggal_pasang, tanggal_error = parse_tanggal(values.get("tanggal_pasang"))
        status = str(values.get("status") or "").strip().upper()

        if not any([nama, kecamatan, nomor_instalasi]):
            continue

        row_errors = []
        if not nama:
            row_errors.append("Nama kosong")
        if not kecamatan:
            row_errors.append("Kecamatan kosong")
        if not nomor_instalasi:
            row_errors.append("Nomor instalasi kosong")
        if tanggal_error:
            row_errors.append(tanggal_error)
        elif not tanggal_pasang:
            row_errors.append("Tanggal pasang kosong")
        if status not in ("SIB", "BK"):
            row_errors.append(f'Status "{status}" tidak valid (harus SIB atau BK)')
        if nomor_instalasi in existing_nomor:
            row_errors.append(f'Nomor instalasi "{nomor_instalasi}" sudah ada di database')

        key = (nama.lower(), kecamatan.lower())
        if key in existing_pelanggan:
            pelanggan_status = f"gabung ke pelanggan lama (id {existing_pelanggan[key]})"
        elif key in batch_pelanggan:
            pelanggan_status = f"gabung ke baris baru lain di file ini (#{batch_pelanggan[key]})"
        else:
            batch_pelanggan[key] = next_batch_no
            pelanggan_status = f"pelanggan baru (#{next_batch_no})"
            next_batch_no += 1

        rows.append({
            "nama": nama, "alamat": str(values.get("alamat") or ""),
            "kelurahan": str(values.get("kelurahan") or ""), "kecamatan": kecamatan,
            "nomor_instalasi": nomor_instalasi, "tanggal_pasang": tanggal_pasang or "",
            "diameter_pipa": str(values.get("diameter_pipa") or ""),
            "tekanan_air": values.get("tekanan_air"),
            "status": status, "petugas": str(values.get("petugas") or ""),
            "keterangan": str(values.get("keterangan") or ""),
            "pelanggan_status": pelanggan_status,
            "errors": row_errors,
        })
        existing_nomor.add(nomor_instalasi)

    return rows


def generate_permohonan_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Permohonan Masuk"

    # 1. Judul Atas Laporan (Baris 2)
    ws.merge_cells("C2:M2")
    title_cell = ws["C2"]
    title_cell.value = "LAPORAN PERMOHONAN MASUK / BUKA KEMBALI DARI HUBUNGAN LANGGANAN"
    title_cell.font = Font(name="Calibri", bold=True, size=12, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2. Struktur Double Header (Baris 4 & 5)
    # Kolom yang tidak pecah akan kita gabung secara vertikal (Row 4 & 5)
    merge_vertikals = ["A4:A5", "B4:B5", "C4:C5", "D4:D5", "E4:E5", "F4:F5", "K4:K5", "L4:L5"]
    for mv in merge_vertikals:
        ws.merge_cells(mv)

    # Gabungan horizontal untuk sub-header Tanggal
    ws.merge_cells("G4:I4") 

    # Isikan nilai teks header sesuai gambar
    ws["A4"] = "No"
    ws["B4"] = "Nama Pelanggan"
    ws["C4"] = "Alamat"
    ws["D4"] = "Kelurahan"
    ws["E4"] = "Kecamatan"
    ws["F4"] = "No.\nSPK"
    
    ws["G4"] = "Tanggal Permohonan Masuk"
    ws["G5"] = "Ke Perencana"
    ws["H5"] = "Survey"
    ws["I5"] = "Kembali Ke\nHublang"
    
    ws["J4"] = "Petugas"
    ws["J5"] = "Survey"
    
    ws["K4"] = "Keterangan"

    # 3. Styling Header (Warna Biru Muda khas Excel Dinas)
    header_fill = PatternFill("solid", start_color="B4C6E7", end_color="B4C6E7")
    header_font = Font(name="Calibri", bold=True, color="000000", size=11)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # Terapkan gaya ke seluruh sel di baris 4 dan 5 (Kolom A sampai K)
    for r in (4, 5):
        ws.row_dimensions[r].height = 24
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    # 4. Lebar Kolom yang Sesuai Dimensi Gambar
    widths = {
        'A': 5,   # No
        'B': 25,  # Nama Pelanggan
        'C': 35,  # Alamat
        'D': 16,  # Kelurahan
        'E': 16,  # Kecamatan
        'F': 8,   # No SPK
        'G': 14,  # Ke Perencana
        'H': 14,  # Survey
        'I': 14,  # Kembali Ke Hublang
        'J': 14,  # Petugas Survey
        'K': 25   # Keterangan
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # 5. Siapkan Baris Kosong Bergrid untuk Pengisian User (Baris 6 sampai 200)
    body_font = Font(name="Calibri", size=11)
    for r in range(6, 201):
        ws.row_dimensions[r].height = 20
        # Format kolom Tanggal agar otomatis rapi saat diisi user
        ws.cell(row=r, column=7).number_format = "DD-MM-YYYY"
        ws.cell(row=r, column=8).number_format = "DD-MM-YYYY"
        ws.cell(row=r, column=9).number_format = "DD-MM-YYYY"
        
        for col in range(1, 12):
            c = ws.cell(row=r, column=col)
            c.font = body_font
            c.border = thin_border

    # Contoh data baris pertama (Baris 6)
    example = [1, "nama", "nama jalan", "kelurahan", "kecamatan", "", "2026-06-02", "2026-06-02", "", "nama petugas", "Keterangan berkas"]
    for col, val in enumerate(example, start=1):
        ws.cell(row=6, column=col, value=val)

    ws.freeze_panes = "A6"
    return wb


def parse_permohonan_upload(fileobj, db):
    wb = load_workbook(fileobj, data_only=True)
    if not wb.worksheets:
        raise ValueError("File Excel ini kosong, tidak ada sheet sama sekali.")
    ws = wb.worksheets[0]

    # Gabungkan teks baris 4 dan baris 5 agar sub-header seperti "Ke Perencana" terbaca sempurna
    col_index = {}
    max_cols = ws.max_column
    
    for col_idx in range(1, max_cols + 1):
        val_r4 = str(ws.cell(row=4, column=col_idx).value or "").strip().lower()
        val_r5 = str(ws.cell(row=5, column=col_idx).value or "").strip().lower()
        combined_header = f"{val_r4} {val_r5}".strip()

        # Cek ke mapping
        for raw_key, field_name in PERMOHONAN_HEADER_MAP.items():
            if raw_key in val_r4 or raw_key in val_r5 or raw_key in combined_header:
                if field_name not in col_index:
                    col_index[field_name] = col_idx - 1  # 0-based index

    # Validasi kolom wajib
    missing = [k for k in ["nama_pelanggan", "kecamatan", "tanggal_permohonan"] if k not in col_index]
    if missing:
        raise ValueError(f"Kolom wajib tidak ditemukan di header file Excel. Mohon pastikan memakai template terbaru SIDAP.")

    rows = []
    # Data dimulai dari baris ke-6
    for row_cells in ws.iter_rows(min_row=6, values_only=False):
        values = {}
        for key, idx in col_index.items():
            values[key] = row_cells[idx].value if idx < len(row_cells) else None

        nama_pelanggan = str(values.get("nama_pelanggan") or "").strip()
        lokasi = str(values.get("lokasi") or "").strip()
        kelurahan = str(values.get("kelurahan") or "").strip()
        kecamatan_raw = str(values.get("kecamatan") or "").strip()
        kecamatan = bersihkan_kecamatan(kecamatan_raw)
        
        # Tentukan jenis dari nama sheet/tab
        sheet_title = ws.title.lower()
        jenis = "BK" if "buka kembali" in sheet_title else "SIB"
        
        tanggal_permohonan, tanggal_permohonan_error = parse_tanggal(values.get("tanggal_permohonan"))
        tanggal_survey, tanggal_survey_error = parse_tanggal(values.get("tanggal_survey"))
        petugas_survey = str(values.get("petugas_survey") or "").strip()
        tanggal_kembali_hublang, tanggal_kembali_error = parse_tanggal(values.get("tanggal_kembali_hublang"))
        keterangan = str(values.get("keterangan") or "").strip()

        # Abaikan baris kosong
        if not any([nama_pelanggan, lokasi, kecamatan, values.get("tanggal_permohonan")]):
            continue

        row_errors = []
        if not nama_pelanggan:
            row_errors.append("Nama pelanggan kosong")
        if not kecamatan:
            row_errors.append("Kecamatan kosong")
        if tanggal_permohonan_error:
            row_errors.append(f"Ke Perencana: {tanggal_permohonan_error}")
        elif not tanggal_permohonan:
            row_errors.append("Tanggal permohonan (Ke Perencana) kosong")
        if tanggal_survey_error:
            row_errors.append(f"Tanggal survey: {tanggal_survey_error}")
        if tanggal_kembali_error:
            row_errors.append(f"Kembali ke Hublang: {tanggal_kembali_error}")

        # Status ditindaklanjuti otomatis
        ditindaklanjuti = None
        if tanggal_survey:
            ditindaklanjuti = 1
        if "selisih" in keterangan.lower() or tanggal_kembali_hublang:
            ditindaklanjuti = 0

        rows.append({
            "nama_pelanggan": nama_pelanggan,
            "lokasi": lokasi,
            "kelurahan": kelurahan,
            "kecamatan": kecamatan,
            "jenis": jenis,
            "tanggal_permohonan": tanggal_permohonan or "",
            "tanggal_survey": tanggal_survey or "",
            "petugas_survey": petugas_survey,
            "ditindaklanjuti": ditindaklanjuti,
            "jenis_pipa": "P.Distribusi" if "distribusi" in keterangan.lower() else "P.Dinas",
            "tanggal_dikirim_hublang": "",
            "tanggal_kembali_hublang": tanggal_kembali_hublang or "",
            "keterangan": keterangan,
            "errors": row_errors,
        })

    return rows


def permohonan_ditindaklanjuti_label(value):
    if value == 1:
        return "Ya"
    if value == 0:
        return "Tidak"
    return "Belum"


@app.context_processor
def inject_helpers():

    db = get_db()
    return dict(
        semua_kecamatan=daftar_kecamatan(db),
        today=datetime.now().strftime("%d %B %Y"),
    )


# ---------------- Halaman Utama Dashboard (Client-Sided) ----------------
@app.route("/")
def dashboard():
    return render_template("dashboard.html")


# ---------------- Endpoint API Data Dashboard (REST API) ----------------
@app.route("/api/dashboard-stats")
def api_dashboard_stats():
    db = get_db()

    total_pelanggan = db.execute("SELECT COUNT(*) c FROM pelanggan").fetchone()["c"]
    total_kecamatan = db.execute("SELECT COUNT(DISTINCT kecamatan) c FROM pelanggan").fetchone()["c"]
    pasang_bulan_ini = db.execute(
        "SELECT COUNT(*) c FROM instalasi WHERE strftime('%Y-%m', tanggal_pasang) = strftime('%Y-%m','now')"
    ).fetchone()["c"]
    jumlah_sib = db.execute("SELECT COUNT(*) c FROM instalasi WHERE status='SIB'").fetchone()["c"]
    jumlah_bk = db.execute("SELECT COUNT(*) c FROM instalasi WHERE status='BK'").fetchone()["c"]

    per_kecamatan_rows = db.execute(
        "SELECT kecamatan, COUNT(*) c FROM pelanggan GROUP BY kecamatan ORDER BY c DESC"
    ).fetchall()
    per_kecamatan = [{"kecamatan": r["kecamatan"], "c": r["c"]} for r in per_kecamatan_rows]

    per_bulan_rows = db.execute(
        """SELECT strftime('%Y-%m', tanggal_pasang) bulan, COUNT(*) c
           FROM instalasi
           GROUP BY bulan ORDER BY bulan ASC LIMIT 6"""
    ).fetchall()
    per_bulan = [{"bulan": r["bulan"], "c": r["c"]} for r in per_bulan_rows]

    recent_rows = db.execute(
        """SELECT p.nama, p.kecamatan, i.nomor_instalasi, i.tanggal_pasang, i.status
           FROM instalasi i JOIN pelanggan p ON p.id = i.pelanggan_id
           ORDER BY i.id DESC LIMIT 5"""
    ).fetchall()
    recent = [{
        "nama": r["nama"],
        "kecamatan": r["kecamatan"],
        "nomor_instalasi": r["nomor_instalasi"],
        "tanggal_pasang": r["tanggal_pasang"],
        "status": r["status"]
    } for r in recent_rows]

    total_permohonan = db.execute("SELECT COUNT(*) c FROM permohonan").fetchone()["c"]
    permohonan_menunggu = db.execute("SELECT COUNT(*) c FROM permohonan WHERE ditindaklanjuti IS NULL").fetchone()["c"]
    permohonan_selisih = db.execute("SELECT COUNT(*) c FROM permohonan WHERE ditindaklanjuti = 0").fetchone()["c"]
    permohonan_selesai = db.execute("SELECT COUNT(*) c FROM permohonan WHERE instalasi_id IS NOT NULL").fetchone()["c"]
    permohonan_bulan_ini = db.execute(
        "SELECT COUNT(*) c FROM permohonan WHERE strftime('%Y-%m', tanggal_permohonan) = strftime('%Y-%m','now')"
    ).fetchone()["c"]

    recent_permohonan_rows = db.execute(
        "SELECT nama_pelanggan, kecamatan, jenis, tanggal_permohonan FROM permohonan ORDER BY id DESC LIMIT 5"
    ).fetchall()
    recent_permohonan = [{
        "nama_pelanggan": r["nama_pelanggan"], "kecamatan": r["kecamatan"],
        "jenis": r["jenis"], "tanggal_permohonan": r["tanggal_permohonan"],
    } for r in recent_permohonan_rows]

    return jsonify({
        "total_pelanggan": total_pelanggan,
        "total_kecamatan": total_kecamatan,
        "pasang_bulan_ini": pasang_bulan_ini,
        "sib_bk": f"{jumlah_sib} / {jumlah_bk}",
        "per_kecamatan": per_kecamatan,
        "per_bulan": per_bulan,
        "recent": recent,
        "total_permohonan": total_permohonan,
        "permohonan_menunggu": permohonan_menunggu,
        "permohonan_selisih": permohonan_selisih,
        "permohonan_selesai": permohonan_selesai,
        "permohonan_bulan_ini": permohonan_bulan_ini,
        "recent_permohonan": recent_permohonan,
    })


# ---------------- Semua Pelanggan (list + cari + filter kecamatan) ----------------
@app.route("/pelanggan/template")
def pelanggan_template():
    wb = generate_template_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="template_data_pelanggan_sidap.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/pelanggan/impor", methods=["GET", "POST"])
def pelanggan_impor():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Pilih file Excel dulu.", "error")
            return redirect(url_for("pelanggan_impor"))
        try:
            rows = parse_upload(file, get_db())
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for("pelanggan_impor"))
        except Exception:
            flash("File tidak bisa dibaca. Pastikan formatnya .xlsx dan pakai template yang disediakan.", "error")
            return redirect(url_for("pelanggan_impor"))

        if not rows:
            flash("Tidak ada baris data yang terbaca dari file ini.", "error")
            return redirect(url_for("pelanggan_impor"))

        os.makedirs(IMPORT_TMP_DIR, exist_ok=True)
        token = uuid.uuid4().hex
        with open(os.path.join(IMPORT_TMP_DIR, f"{token}.json"), "w") as f:
            json.dump(rows, f)

        jumlah_error = sum(1 for r in rows if r["errors"])
        jumlah_baru = len({r["pelanggan_status"] for r in rows if "pelanggan baru" in r["pelanggan_status"]})
        return render_template(
            "import_preview.html", rows=rows, token=token,
            jumlah_error=jumlah_error, jumlah_baru=jumlah_baru, total=len(rows),
        )

    return render_template("import_upload.html")


@app.route("/pelanggan/impor/konfirmasi", methods=["POST"])
def pelanggan_impor_konfirmasi():
    token = request.form.get("token", "")
    path = os.path.join(IMPORT_TMP_DIR, f"{token}.json")
    if not os.path.exists(path):
        flash("Sesi impor sudah kedaluwarsa, silakan upload ulang.", "error")
        return redirect(url_for("pelanggan_impor"))

    with open(path) as f:
        rows = json.load(f)
    os.remove(path)

    valid_rows = [r for r in rows if not r["errors"]]
    if not valid_rows:
        flash("Tidak ada baris valid yang bisa disimpan -- semua baris punya error.", "error")
        return redirect(url_for("pelanggan_impor"))

    db = get_db()
    pelanggan_cache = {}
    for r in db.execute("SELECT id, nama, kecamatan FROM pelanggan").fetchall():
        pelanggan_cache[(r["nama"].strip().lower(), r["kecamatan"].strip().lower())] = r["id"]

    tersimpan = 0
    for r in valid_rows:
        key = (r["nama"].strip().lower(), r["kecamatan"].strip().lower())
        if key not in pelanggan_cache:
            cur = db.execute(
                "INSERT INTO pelanggan (nama, alamat, kelurahan, kecamatan) VALUES (?,?,?,?)",
                (r["nama"], r["alamat"], r["kelurahan"], r["kecamatan"]),
            )
            pelanggan_cache[key] = cur.lastrowid

        db.execute(
            """INSERT INTO instalasi
               (pelanggan_id, nomor_instalasi, tanggal_pasang, diameter_pipa,
                tekanan_air, status, petugas, keterangan)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                pelanggan_cache[key], r["nomor_instalasi"], r["tanggal_pasang"],
                r["diameter_pipa"], r["tekanan_air"] or None, r["status"],
                r["petugas"] or None, r["keterangan"],
            ),
        )
        tersimpan += 1

    db.commit()
    dilewati = len(rows) - tersimpan
    pesan = f"{tersimpan} baris berhasil diimpor."
    if dilewati:
        pesan += f" {dilewati} baris dilewati karena ada error."
    flash(pesan)
    return redirect(url_for("pelanggan_list"))


@app.route("/permohonan/template")
def permohonan_template():
    wb = generate_permohonan_template_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="template_data_permohonan_sidap.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/permohonan/impor", methods=["GET", "POST"])
def permohonan_impor():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Pilih file Excel dulu.", "error")
            return redirect(url_for("permohonan_impor"))
        try:
            rows = parse_permohonan_upload(file, get_db())
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for("permohonan_impor"))
        except Exception:
            flash("File tidak bisa dibaca. Pastikan formatnya .xlsx dan pakai template yang disediakan.", "error")
            return redirect(url_for("permohonan_impor"))

        if not rows:
            flash("Tidak ada baris data yang terbaca dari file ini.", "error")
            return redirect(url_for("permohonan_impor"))

        os.makedirs(IMPORT_TMP_DIR, exist_ok=True)
        token = uuid.uuid4().hex
        with open(os.path.join(IMPORT_TMP_DIR, f"{token}.json"), "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False)

        jumlah_error = sum(1 for r in rows if r["errors"])
        jumlah_valid = len(rows) - jumlah_error
        return render_template(
            "permohonan_import_preview.html", rows=rows, token=token,
            jumlah_error=jumlah_error, jumlah_valid=jumlah_valid, total=len(rows),
        )

    return render_template("permohonan_import_upload.html")


@app.route("/permohonan/impor/konfirmasi", methods=["POST"])
def permohonan_impor_konfirmasi():
    token = request.form.get("token", "")
    path = os.path.join(IMPORT_TMP_DIR, f"{token}.json")
    if not os.path.exists(path):
        flash("Sesi impor sudah kedaluwarsa, silakan upload ulang.", "error")
        return redirect(url_for("permohonan_impor"))

    with open(path, encoding="utf-8") as f:
        rows = json.load(f)
    
    if os.path.exists(path):
        os.remove(path)

    valid_rows = [r for r in rows if not r.get("errors")]
    if not valid_rows:
        flash("Tidak ada baris valid yang bisa disimpan -- semua baris punya error.", "error")
        return redirect(url_for("permohonan_impor"))

    db = get_db()
    
    # 1. Pastikan kolom no_spk ada di tabel permohonan (Auto-Migration sederhana)
    try:
        db.execute("ALTER TABLE permohonan ADD COLUMN no_spk TEXT")
        db.commit()
    except sqlite3.OperationalError:
        pass  # Abaikan jika kolom no_spk sudah ada sebelumnya

    # 2. Ambil nilai No. SPK terbesar saat ini untuk Auto-Increment
    max_spk_row = db.execute("SELECT MAX(CAST(no_spk AS INTEGER)) m FROM permohonan").fetchone()
    current_spk = int(max_spk_row["m"]) if max_spk_row and max_spk_row["m"] else 0

    tersimpan = 0
    for r in valid_rows:
        current_spk += 1
        
        db.execute(
            """INSERT INTO permohonan
               (nama_pelanggan, lokasi, kelurahan, kecamatan, jenis, no_spk, tanggal_permohonan,
                tanggal_survey, petugas_survey, ditindaklanjuti, jenis_pipa,
                tanggal_dikirim_hublang, tanggal_kembali_hublang, keterangan)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                r.get("nama_pelanggan", ""),
                r.get("lokasi", ""),
                r.get("kelurahan", ""),
                r.get("kecamatan", ""),
                r.get("jenis", "SIB"),
                str(current_spk),
                r.get("tanggal_permohonan", ""),
                r.get("tanggal_survey") or None,
                r.get("petugas_survey") or None,
                r.get("ditindaklanjuti"),
                r.get("jenis_pipa") or None,
                r.get("tanggal_dikirim_hublang") or None,
                r.get("tanggal_kembali_hublang") or None,
                r.get("keterangan", ""),
            ),
        )
        tersimpan += 1
        
    db.commit()

    dilewati = len(rows) - tersimpan
    pesan = f"{tersimpan} baris permohonan berhasil diimpor dengan No. SPK otomatis."
    if dilewati:
        pesan += f" {dilewati} baris dilewati karena ada error."
    flash(pesan)
    return redirect(url_for("permohonan_list"))


# ---------------- Cari Pelanggan ----------------
@app.route("/cari")
def cari_pelanggan():

    # Server hanya melempar halaman HTML kosong tanpa memproses database
    return render_template("cari_pelanggan.html")


@app.route("/api/cari")
def api_cari_pelanggan():
    db = get_db()
    q = request.args.get("q", "").strip()
    scope = request.args.get("scope", "semua")
    rows = []

    if q:
        query = """SELECT p.id pelanggan_id, p.nama, p.kecamatan,
                          i.id instalasi_id, i.nomor_instalasi, i.tanggal_pasang, i.status
                   FROM instalasi i JOIN pelanggan p ON p.id = i.pelanggan_id
                   WHERE 1=1"""
        params = []
        if scope == "nama":
            query += " AND p.nama LIKE ?"
            params.append(f"%{q}%")
        elif scope == "instalasi":
            query += " AND i.nomor_instalasi LIKE ?"
            params.append(f"%{q}%")
        elif scope == "kecamatan":
            query += " AND p.kecamatan LIKE ?"
            params.append(f"%{q}%")
        else:
            query += " AND (p.nama LIKE ? OR i.nomor_instalasi LIKE ? OR p.kecamatan LIKE ?)"
            params += [f"%{q}%", f"%{q}%", f"%{q}%"]
        query += " ORDER BY p.nama, i.tanggal_pasang"

        found = db.execute(query, params).fetchall()
        seq_counter = {}
        for r in found:
            seq_counter[r["pelanggan_id"]] = seq_counter.get(r["pelanggan_id"], 0) + 1
            d = dict(r)
            d["urutan_instalasi"] = seq_counter[r["pelanggan_id"]]
            rows.append(d)

    # Mengembalikan hasil pencarian dalam bentuk JSON mentah
    return jsonify({
        "q": q,
        "scope": scope,
        "results": rows
    })

# 1. Rute untuk menampilkan halaman (Ini yang membuat halaman Semua Pelanggan terbuka)
@app.route("/pelanggan")
def pelanggan_list():
    return render_template("pelanggan_list.html")


# 2. Rute API untuk menyuplai data (INI YANG HILANG ATAU BELUM TERSIMPAN)
@app.route("/api/pelanggan")
def api_pelanggan_list():
    db = get_db()
    q = request.args.get("q", "").strip()
    kecamatan = request.args.get("kecamatan", "semua")
    status = request.args.get("status", "semua")
    tahun = request.args.get("tahun", "").strip()
    bulan = request.args.get("bulan", "").strip()
    page = max(1, request.args.get("page", 1, type=int))
    per_page = 30

    # v0.9.5: list sekarang satu baris = satu instalasi (bukan satu baris = satu
    # pelanggan yang diringkas), sesuai keputusan revisi UI -- pelanggan dengan
    # banyak instalasi (mis. kasus Sarpita) muncul sebagai beberapa baris.
    # Pelanggan yang belum punya instalasi sama sekali tetap muncul 1 baris
    # (kolom instalasi kosong) lewat LEFT JOIN.
    conditions = []
    params = []
    if q:
        conditions.append("(p.nama LIKE ? OR i.nomor_instalasi LIKE ?)")
        params += [f"%{q}%", f"%{q}%"]
    if kecamatan != "semua":
        conditions.append("p.kecamatan = ?")
        params.append(kecamatan)
    if status != "semua":
        conditions.append("i.status = ?")
        params.append(status)
    if tahun:
        conditions.append("strftime('%Y', i.tanggal_pasang) = ?")
        params.append(tahun)
    if bulan:
        conditions.append("strftime('%m', i.tanggal_pasang) = ?")
        params.append(bulan.zfill(2))

    where_clause = " AND ".join(conditions) if conditions else "1=1"

    query = f"""SELECT p.id pelanggan_id, p.nama, p.alamat, p.kelurahan, p.kecamatan,
                       i.id instalasi_id, i.nomor_instalasi, i.tanggal_pasang,
                       i.diameter_pipa, i.tekanan_air, i.status, i.petugas, i.keterangan
                FROM pelanggan p
                LEFT JOIN instalasi i ON i.pelanggan_id = p.id
                WHERE {where_clause}
                ORDER BY p.nama, i.tanggal_pasang"""

    all_rows = db.execute(query, params).fetchall()
    total = len(all_rows)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)

    sliced_rows = all_rows[(page - 1) * per_page: page * per_page]

    rows = [{
        "pelanggan_id": r["pelanggan_id"],
        "nama": r["nama"],
        "alamat": r["alamat"],
        "kelurahan": r["kelurahan"],
        "kecamatan": r["kecamatan"],
        "instalasi_id": r["instalasi_id"],
        "nomor_instalasi": r["nomor_instalasi"] or "",
        "tanggal_pasang": r["tanggal_pasang"] or "",
        "diameter_pipa": r["diameter_pipa"] or "",
        "tekanan_air": r["tekanan_air"] if r["tekanan_air"] is not None else "",
        "status": r["status"] or "",
        "petugas": r["petugas"] or "",
        "keterangan": r["keterangan"] or "",
    } for r in sliced_rows]

    kecamatan_list = daftar_kecamatan(db)

    return jsonify({
        "rows": rows,
        "total": total,
        "page": page,
        "total_pages": total_pages,
        "semua_kecamatan": kecamatan_list
    })


@app.route("/pelanggan/export")
def pelanggan_export():
    db = get_db()
    q = request.args.get("q", "").strip()
    kecamatan = request.args.get("kecamatan", "semua")
    status = request.args.get("status", "semua")
    tahun = request.args.get("tahun", "").strip()
    bulan = request.args.get("bulan", "").strip()

    conditions = []
    params = []
    if q:
        conditions.append("(p.nama LIKE ? OR i.nomor_instalasi LIKE ?)")
        params += [f"%{q}%", f"%{q}%"]
    if kecamatan != "semua":
        conditions.append("p.kecamatan = ?")
        params.append(kecamatan)
    if status != "semua":
        conditions.append("i.status = ?")
        params.append(status)
    if tahun:
        conditions.append("strftime('%Y', i.tanggal_pasang) = ?")
        params.append(tahun)
    if bulan:
        conditions.append("strftime('%m', i.tanggal_pasang) = ?")
        params.append(bulan.zfill(2))

    where_clause = " AND ".join(conditions) if conditions else "1=1"
    rows = db.execute(
        f"""SELECT p.nama, p.alamat, p.kelurahan, p.kecamatan,
                   i.nomor_instalasi, i.tanggal_pasang, i.status, i.petugas, i.keterangan
            FROM pelanggan p
            LEFT JOIN instalasi i ON i.pelanggan_id = p.id
            WHERE {where_clause}
            ORDER BY p.nama, i.tanggal_pasang""",
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Semua Pelanggan"
    ws.append([
        "Nama", "Alamat", "Kelurahan", "Kecamatan",
        "No. Instalasi", "Tanggal Pasang", "Status", "Petugas", "Keterangan"
    ])
    for r in rows:
        ws.append([
            r["nama"], r["alamat"], r["kelurahan"], r["kecamatan"],
            r["nomor_instalasi"] or "", r["tanggal_pasang"] or "",
            r["status"] or "", r["petugas"] or "", r["keterangan"] or "",
        ])

    for idx, width in enumerate([25, 35, 20, 20, 16, 15, 10, 15, 30], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"data_pelanggan_{tahun or 'semua'}_{bulan or 'semua'}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ---------------- Detail satu pelanggan ----------------
@app.route("/pelanggan/<int:pid>")
def pelanggan_detail(pid):
    db = get_db()
    pelanggan = db.execute("SELECT * FROM pelanggan WHERE id=?", (pid,)).fetchone()
    instalasi = db.execute(
        "SELECT * FROM instalasi WHERE pelanggan_id=? ORDER BY tanggal_pasang", (pid,)
    ).fetchall()
    return render_template("pelanggan_detail.html", pelanggan=pelanggan, instalasi=instalasi)


# ---------------- Tambah pelanggan + instalasi ----------------
@app.route("/pelanggan/tambah", methods=["GET", "POST"])
def pelanggan_tambah():
    db = get_db()
    if request.method == "POST":
        nomor_instalasi = request.form["nomor_instalasi"].strip()
        pelanggan_existing_id = request.form.get("pelanggan_existing_id", "").strip()

        dup = db.execute(
            "SELECT 1 FROM instalasi WHERE nomor_instalasi=?", (nomor_instalasi,)
        ).fetchone()
        if dup:
            flash(f"Nomor instalasi \"{nomor_instalasi}\" sudah dipakai instalasi lain. Cek kembali sebelum simpan.", "error")
            return render_template("pelanggan_form.html", pelanggan=None, form_data=request.form,
                                    semua_pelanggan_pilihan=daftar_pelanggan_pilihan(db))

        if not is_valid_iso_date(request.form.get("tanggal_pasang", "")):
            flash("Tanggal pasang tidak valid. Pakai date picker, jangan ketik manual.", "error")
            return render_template("pelanggan_form.html", pelanggan=None, form_data=request.form,
                                    semua_pelanggan_pilihan=daftar_pelanggan_pilihan(db))

        if pelanggan_existing_id:
            existing = db.execute("SELECT id, nama FROM pelanggan WHERE id=?", (pelanggan_existing_id,)).fetchone()
            if not existing:
                flash("Pelanggan yang dipilih tidak ditemukan. Pilih dari daftar, jangan ketik manual.", "error")
                return render_template("pelanggan_form.html", pelanggan=None, form_data=request.form,
                                        semua_pelanggan_pilihan=daftar_pelanggan_pilihan(db))
            pelanggan_id = existing["id"]
            nama_untuk_pesan = existing["nama"]
        else:
            cur = db.execute(
                "INSERT INTO pelanggan (nama, alamat, kelurahan, kecamatan) VALUES (?,?,?,?)",
                (
                    request.form["nama"],
                    request.form.get("alamat", ""),
                    request.form.get("kelurahan", ""),
                    request.form["kecamatan"],
                ),
            )
            pelanggan_id = cur.lastrowid
            nama_untuk_pesan = request.form["nama"]

        db.execute(
            """INSERT INTO instalasi
               (pelanggan_id, nomor_instalasi, tanggal_pasang, diameter_pipa,
                tekanan_air, status, petugas, keterangan)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                pelanggan_id,
                nomor_instalasi,
                request.form["tanggal_pasang"],
                request.form.get("diameter_pipa", ""),
                request.form.get("tekanan_air") or None,
                request.form["status"],
                request.form.get("petugas") or None,
                request.form.get("keterangan", ""),
            ),
        )
        db.commit()
        flash(f"Instalasi {nomor_instalasi} tersimpan untuk {nama_untuk_pesan}.")
        return redirect(url_for("pelanggan_list"))

    return render_template("pelanggan_form.html", pelanggan=None, form_data=None,
                            semua_pelanggan_pilihan=daftar_pelanggan_pilihan(db))


# ---------------- Tambah instalasi baru ----------------
@app.route("/pelanggan/<int:pid>/instalasi/tambah", methods=["GET", "POST"])
def instalasi_tambah(pid):
    db = get_db()
    pelanggan = db.execute("SELECT * FROM pelanggan WHERE id=?", (pid,)).fetchone()
    if request.method == "POST":
        nomor_instalasi = request.form["nomor_instalasi"].strip()
        dup = db.execute(
            "SELECT 1 FROM instalasi WHERE nomor_instalasi=?", (nomor_instalasi,)
        ).fetchone()
        if dup:
            flash(f"Nomor instalasi \"{nomor_instalasi}\" sudah dipakai instalasi lain. Cek kembali sebelum simpan.", "error")
            return render_template("instalasi_form.html", pelanggan=pelanggan, form_data=request.form)

        if not is_valid_iso_date(request.form.get("tanggal_pasang", "")):
            flash("Tanggal pasang tidak valid. Pakai date picker, jangan ketik manual.", "error")
            return render_template("instalasi_form.html", pelanggan=pelanggan, form_data=request.form)

        db.execute(
            """INSERT INTO instalasi
               (pelanggan_id, nomor_instalasi, tanggal_pasang, diameter_pipa,
                tekanan_air, status, petugas, keterangan)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                pid,
                nomor_instalasi,
                request.form["tanggal_pasang"],
                request.form.get("diameter_pipa", ""),
                request.form.get("tekanan_air") or None,
                request.form["status"],
                request.form.get("petugas") or None,
                request.form.get("keterangan", ""),
            ),
        )
        db.commit()
        flash(f"Instalasi baru untuk {pelanggan['nama']} tersimpan.")
        return redirect(url_for("pelanggan_detail", pid=pid))

    return render_template("instalasi_form.html", pelanggan=pelanggan, form_data=None)


# ---------------- Edit / hapus instalasi + data pelanggan ----------------
@app.route("/instalasi/<int:iid>/edit", methods=["GET", "POST"])
def instalasi_edit(iid):
    db = get_db()
    inst = db.execute("SELECT * FROM instalasi WHERE id=?", (iid,)).fetchone()
    if not inst:
        flash("Instalasi tidak ditemukan.", "error")
        return redirect(url_for("pelanggan_list"))
        
    pelanggan = db.execute("SELECT * FROM pelanggan WHERE id=?", (inst["pelanggan_id"],)).fetchone()

    if request.method == "POST":
        nomor_instalasi = request.form["nomor_instalasi"].strip()
        dup = db.execute(
            "SELECT 1 FROM instalasi WHERE nomor_instalasi=? AND id!=?", (nomor_instalasi, iid)
        ).fetchone()
        if dup:
            flash(f'Nomor instalasi "{nomor_instalasi}" sudah dipakai instalasi lain. Cek kembali sebelum simpan.', "error")
            return render_template("instalasi_edit.html", instalasi=inst, pelanggan=pelanggan)

        if not is_valid_iso_date(request.form.get("tanggal_pasang", "")):
            flash("Tanggal pasang tidak valid. Pakai date picker, jangan ketik manual.", "error")
            return render_template("instalasi_edit.html", instalasi=inst, pelanggan=pelanggan)

        # 1. Update Data Pelanggan (Nama, Alamat, Kelurahan, Kecamatan)
        db.execute(
            "UPDATE pelanggan SET nama=?, alamat=?, kelurahan=?, kecamatan=? WHERE id=?",
            (
                request.form.get("nama", pelanggan["nama"]),
                request.form.get("alamat", ""),
                request.form.get("kelurahan", ""),
                request.form.get("kecamatan", pelanggan["kecamatan"]),
                inst["pelanggan_id"],
            ),
        )

        # 2. Update Data Instalasi
        db.execute(
            """UPDATE instalasi SET nomor_instalasi=?, tanggal_pasang=?, diameter_pipa=?,
               tekanan_air=?, status=?, petugas=?, keterangan=? WHERE id=?""",
            (
                nomor_instalasi,
                request.form["tanggal_pasang"],
                request.form.get("diameter_pipa", ""),
                request.form.get("tekanan_air") or None,
                request.form["status"],
                request.form.get("petugas") or None,
                request.form.get("keterangan", ""),
                iid,
            ),
        )
        db.commit()
        flash("Data pelanggan & instalasi berhasil diperbarui.")
        # Mengarahkan kembali ke daftar Semua Pelanggan
        return redirect(url_for("pelanggan_list"))

    return render_template("instalasi_edit.html", instalasi=inst, pelanggan=pelanggan)


@app.route("/instalasi/<int:iid>/hapus", methods=["POST"])
def instalasi_hapus(iid):
    db = get_db()
    db.execute("DELETE FROM instalasi WHERE id=?", (iid,))
    db.commit()
    flash("Instalasi berhasil dihapus.")
    # Mengarahkan kembali ke daftar Semua Pelanggan
    return redirect(url_for("pelanggan_list"))


# ---------------- Menu Kecamatan ----------------
@app.route("/kecamatan")
def kecamatan_list():
    db = get_db()
    rows = db.execute(
        """SELECT p.kecamatan,
                  COUNT(DISTINCT p.id) jumlah_pelanggan,
                  SUM(CASE WHEN i.status='SIB' THEN 1 ELSE 0 END) jumlah_sib,
                  SUM(CASE WHEN i.status='BK' THEN 1 ELSE 0 END) jumlah_bk
           FROM pelanggan p LEFT JOIN instalasi i ON i.pelanggan_id = p.id
           GROUP BY p.kecamatan ORDER BY p.kecamatan"""
    ).fetchall()
    return render_template("kecamatan_list.html", rows=rows)

@app.route("/api/kecamatan-stats")
def api_kecamatan_stats():
    db = get_db()
    rows = db.execute(
        """SELECT p.kecamatan,
                  COUNT(DISTINCT p.id) jumlah_pelanggan,
                  SUM(CASE WHEN i.status='SIB' THEN 1 ELSE 0 END) jumlah_sib,
                  SUM(CASE WHEN i.status='BK' THEN 1 ELSE 0 END) jumlah_bk
           FROM pelanggan p LEFT JOIN instalasi i ON i.pelanggan_id = p.id
           GROUP BY p.kecamatan ORDER BY p.kecamatan"""
    ).fetchall()
    
    data = [{
        "kecamatan": r["kecamatan"],
        "jumlah_pelanggan": r["jumlah_pelanggan"],
        "jumlah_sib": r["jumlah_sib"] or 0,
        "jumlah_bk": r["jumlah_bk"] or 0
    } for r in rows]
    
    return jsonify(data)

# ---------------- Menu Tahun & Bulan ----------------
# ---------------- Menu Tahun & Bulan (REST API & Client-Sided) ----------------
@app.route("/periode")
def periode_list():
    return render_template("periode_list.html")


@app.route("/api/periode-stats")
def api_periode_stats():
    db = get_db()
    
    # 1. Ambil daftar tahun unik untuk dropdown
    tahun_list = [
        r["t"] for r in db.execute(
            """SELECT DISTINCT strftime('%Y', tanggal_pasang) t FROM instalasi
               WHERE tanggal_pasang IS NOT NULL AND strftime('%Y', tanggal_pasang) IS NOT NULL
               ORDER BY t DESC"""
        ).fetchall()
    ]
    
    # Tentukan tahun terpilih (default tahun terbaru)
    tahun_terpilih = request.args.get("tahun", "") or (tahun_list[0] if tahun_list else datetime.now().strftime("%Y"))
    if tahun_list and tahun_terpilih not in tahun_list:
        tahun_terpilih = tahun_list[0]

    # 2. Ambil data pemasangan per bulan pada tahun tersebut
    per_bulan = db.execute(
        """SELECT strftime('%m', tanggal_pasang) bulan, COUNT(*) jumlah
           FROM instalasi
           WHERE strftime('%Y', tanggal_pasang) = ?
           GROUP BY bulan""",
        (tahun_terpilih,),
    ).fetchall()
    
    bulan_map = {r["bulan"]: r["jumlah"] for r in per_bulan}
    nama_bulan = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
    
    ringkasan = [
        {"nomor": f"{i + 1:02d}", "nama": nama, "jumlah": bulan_map.get(f"{i + 1:02d}", 0)}
        for i, nama in enumerate(nama_bulan)
    ]
    max_jumlah = max([r["jumlah"] for r in ringkasan], default=1) or 1

    # 3. Hitung jumlah data yang tanggal pasangnya rusak/kosong
    rusak = db.execute(
        """SELECT COUNT(*) c FROM instalasi
           WHERE tanggal_pasang IS NULL OR strftime('%Y', tanggal_pasang) IS NULL"""
    ).fetchone()["c"]

    return jsonify({
        "tahun_list": tahun_list,
        "tahun_terpilih": tahun_terpilih,
        "ringkasan": ringkasan,
        "max_jumlah": max_jumlah,
        "rusak": rusak
    })


# ---------------- Rekapitulasi Laporan Bulanan ----------------
NAMA_BULAN = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
              "Juli", "Agustus", "September", "Oktober", "November", "Desember"]


def hitung_rekap_bulanan(db, tahun, bulan):
    semua_kec = daftar_kecamatan(db)
    data = db.execute(
        """SELECT p.kecamatan,
                  SUM(CASE WHEN i.status='SIB' THEN 1 ELSE 0 END) sib,
                  SUM(CASE WHEN i.status='BK' THEN 1 ELSE 0 END) bk
           FROM instalasi i JOIN pelanggan p ON p.id = i.pelanggan_id
           WHERE strftime('%Y', i.tanggal_pasang) = ? AND strftime('%m', i.tanggal_pasang) = ?
           GROUP BY p.kecamatan""",
        (tahun, bulan),
    ).fetchall()
    peta = {r["kecamatan"]: {"sib": r["sib"] or 0, "bk": r["bk"] or 0} for r in data}

    rows = []
    total_sib = total_bk = 0
    for kec in semua_kec:
        sib = peta.get(kec, {}).get("sib", 0)
        bk = peta.get(kec, {}).get("bk", 0)
        rows.append({"kecamatan": kec, "sib": sib, "bk": bk, "jumlah": sib + bk})
        total_sib += sib
        total_bk += bk

    return rows, total_sib, total_bk


@app.route("/laporan")
def laporan_bulanan():
    # Hanya melempar halaman kosong agar browser yang menyusun tabelnya
    return render_template("laporan_bulanan.html")

@app.route("/laporan/unduh")
def laporan_unduh_xlsx():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    db = get_db()
    tahun = request.args.get("tahun", "").strip()
    bulan = request.args.get("bulan", "").strip()
    
    # 1. Ambil data rekap
    rows, total_sib, total_bk = hitung_rekap_bulanan(db, tahun, bulan)
    nama_bulan = NAMA_BULAN[int(bulan) - 1] if bulan else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Bulanan"

    # Define Styling
    font_title = Font(name="Calibri", bold=True, size=12)
    font_sub = Font(name="Calibri", bold=True, size=10)
    font_header = Font(name="Calibri", bold=True, size=10)
    font_body = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", bold=True, size=10)

    fill_header = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_side = Side(style="thin", color="000000")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 2. Baris Judul Laporan (Merge A1:E1 dan A2:E2)
    ws.merge_cells("A1:E1")
    ws["A1"] = "LAPORAN BULANAN"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:E2")
    ws["A2"] = f"BULAN: {nama_bulan.upper()} {tahun}"
    ws["A2"].font = font_sub
    ws["A2"].alignment = align_center

    # Baris 3 kosong untuk jarak

    # 3. Header Tabel (Baris 4)
    headers = ["NO", "KECAMATAN", "SIB", "BK", "JUMLAH"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
    ws.row_dimensions[4].height = 24

    # 4. Isi Data Kecamatan (Baris 5 dst)
    current_row = 5
    for idx, r in enumerate(rows, start=1):
        # Kolom 1: NO
        c_no = ws.cell(row=current_row, column=1, value=idx)
        c_no.alignment = align_center
        
        # Kolom 2: KECAMATAN
        c_kec = ws.cell(row=current_row, column=2, value=r["kecamatan"].upper())
        c_kec.alignment = align_left
        
        # Kolom 3: SIB
        c_sib = ws.cell(row=current_row, column=3, value=r["sib"] if r["sib"] > 0 else "-")
        c_sib.alignment = align_center
        
        # Kolom 4: BK
        c_bk = ws.cell(row=current_row, column=4, value=r["bk"] if r["bk"] > 0 else "-")
        c_bk.alignment = align_center
        
        # Kolom 5: JUMLAH
        c_jml = ws.cell(row=current_row, column=5, value=r["jumlah"])
        c_jml.alignment = align_center

        # Terapkan font & border untuk semua sel di baris ini
        for col_idx in range(1, 6):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_body
            cell.border = border_all
        
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # 5. Baris Total Paling Bawah
    ws.merge_cells(f"A{current_row}:B{current_row}")
    ws.cell(row=current_row, column=1, value="JUMLAH").alignment = align_center
    ws.cell(row=current_row, column=3, value=total_sib).alignment = align_center
    ws.cell(row=current_row, column=4, value=total_bk).alignment = align_center
    ws.cell(row=current_row, column=5, value=total_sib + total_bk).alignment = align_center

    for col_idx in range(1, 6):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = font_bold
        cell.border = border_all

    ws.row_dimensions[current_row].height = 22

    # 6. Atur Lebar Kolom
    col_widths = {
        'A': 8,   # NO
        'B': 25,  # KECAMATAN
        'C': 15,  # SIB
        'D': 15,  # BK
        'E': 15   # JUMLAH
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 7. Simpan file ke BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Laporan_Bulanan_{tahun}_{bulan}.xlsx"
    )

@app.route("/api/laporan-stats")
def api_laporan_stats():
    db = get_db()
    
    # 1. Ambil daftar tahun unik untuk dropdown filter
    tahun_list = [
        r["t"] for r in db.execute(
            """SELECT DISTINCT strftime('%Y', tanggal_pasang) t FROM instalasi
               WHERE tanggal_pasang IS NOT NULL AND strftime('%Y', tanggal_pasang) IS NOT NULL
               ORDER BY t DESC"""
        ).fetchall()
    ]
    tahun = request.args.get("tahun", "") or (tahun_list[0] if tahun_list else datetime.now().strftime("%Y"))
    bulan = request.args.get("bulan", "") or datetime.now().strftime("%m")

    # 2. Hitung rekap bulanan menggunakan fungsi helper bawaan
    rows, total_sib, total_bk = hitung_rekap_bulanan(db, tahun, bulan)
    nama_bulan = NAMA_BULAN[int(bulan) - 1]

    return jsonify({
        "tahun_list": tahun_list,
        "tahun_terpilih": tahun,
        "bulan_terpilih": bulan,
        "nama_bulan": nama_bulan,
        "nama_bulan_list": list(enumerate(NAMA_BULAN, start=1)),
        "rows": rows,
        "total_sib": total_sib,
        "total_bk": total_bk,
        "total_jumlah": total_sib + total_bk
    })


# ---------------- Eksekusi Aplikasi Utama ----------------
KETERANGAN_PERMOHONAN = [
    "Berkas / Diproses",
    "Permohonan Double",
    "Meter Air Sudah Terpasang",
    "Rumah Tidak Ditemukan",
    "Butuh Pipa Distribusi",
    "Tidak ada Akses Jalan Pipa",
    "Alamat tidak dapat ditemukan",
    "Rumah Tanah PJKA",
    "Bekas Pemutusan / Buka Kembali",
]


def panjang_pipa_label(jenis_pipa):
    """Panjang pipa itu bukan field terpisah -- diturunkan dari jenis_pipa.
    P.Dinas selalu <=10 meter, P.Distribusi selalu >10 meter (sesuai
    konfirmasi kabag)."""
    if jenis_pipa == "P.Dinas":
        return "\u226410 meter"
    if jenis_pipa == "P.Distribusi":
        return ">10 meter (P.Distribusi)"
    return None


def cari_atau_buat_pelanggan(db, nama, alamat, kelurahan, kecamatan):
    """Cocokin ke pelanggan existing (nama+ALAMAT sama persis, bukan cuma
    nama+kecamatan) -- biar gak salah gabung. Kasus nyata dari laporan
    Hublang: satu nama bisa muncul berkali-kali di kecamatan yang sama tapi
    alamat beda-beda (developer ngajuin permohonan buat banyak unit rumah
    sekaligus) -- itu HARUS dianggap pelanggan/properti berbeda, bukan satu
    pelanggan yang kebetulan namanya sama."""
    existing = db.execute(
        "SELECT id FROM pelanggan WHERE LOWER(nama)=LOWER(?) AND LOWER(alamat)=LOWER(?) AND LOWER(kecamatan)=LOWER(?)",
        (nama.strip(), alamat.strip(), kecamatan.strip()),
    ).fetchone()
    if existing:
        return existing["id"]
    cur = db.execute(
        "INSERT INTO pelanggan (nama, alamat, kelurahan, kecamatan) VALUES (?,?,?,?)",
        (nama, alamat, kelurahan, kecamatan),
    )
    return cur.lastrowid


# ---------------- Permohonan (dari Hublang) ----------------
@app.route("/permohonan")
def permohonan_list():
    db = get_db()
    jenis = request.args.get("jenis", "semua")
    status = request.args.get("status", "semua")  # semua/menunggu/ditindaklanjuti/selisih/selesai
    kecamatan = request.args.get("kecamatan", "semua")
    tahun = request.args.get("tahun", "").strip()
    bulan = request.args.get("bulan", "").strip()
    q = request.args.get("q", "").strip()
    page = max(1, request.args.get("page", 1, type=int))
    per_page = 30

    conditions = []
    params = []
    if jenis != "semua":
        conditions.append("jenis = ?")
        params.append(jenis)
    if kecamatan != "semua":
        conditions.append("kecamatan = ?")
        params.append(kecamatan)
    if status == "menunggu":
        conditions.append("ditindaklanjuti IS NULL")
    elif status == "ditindaklanjuti":
        conditions.append("ditindaklanjuti = 1 AND instalasi_id IS NULL")
    elif status == "selisih":
        conditions.append("ditindaklanjuti = 0")
    elif status == "selesai":
        conditions.append("instalasi_id IS NOT NULL")
    if tahun:
        conditions.append("strftime('%Y', tanggal_permohonan) = ?")
        params.append(tahun)
    if bulan:
        conditions.append("strftime('%m', tanggal_permohonan) = ?")
        params.append(bulan.zfill(2))
    if q:
        conditions.append("(nama_pelanggan LIKE ? OR lokasi LIKE ? OR kecamatan LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]

    where_clause = " AND ".join(conditions) if conditions else "1=1"
    rows_all = db.execute(
        f"SELECT * FROM permohonan WHERE {where_clause} ORDER BY tanggal_permohonan DESC, id DESC",
        params,
    ).fetchall()

    total = len(rows_all)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    rows = rows_all[(page - 1) * per_page: page * per_page]

    ringkas = {
        "menunggu": db.execute("SELECT COUNT(*) c FROM permohonan WHERE ditindaklanjuti IS NULL").fetchone()["c"],
        "ditindaklanjuti": db.execute("SELECT COUNT(*) c FROM permohonan WHERE ditindaklanjuti = 1 AND instalasi_id IS NULL").fetchone()["c"],
        "selisih": db.execute("SELECT COUNT(*) c FROM permohonan WHERE ditindaklanjuti = 0").fetchone()["c"],
        "selesai": db.execute("SELECT COUNT(*) c FROM permohonan WHERE instalasi_id IS NOT NULL").fetchone()["c"],
    }

    return render_template(
        "permohonan_list.html", rows=rows, jenis=jenis, status=status, kecamatan=kecamatan,
        tahun=tahun, bulan=bulan, q=q,
        page=page, total_pages=total_pages, total=total, ringkas=ringkas,
    )


@app.route("/permohonan/export")
def permohonan_export():
    db = get_db()
    jenis = request.args.get("jenis", "semua")
    status = request.args.get("status", "semua")
    kecamatan = request.args.get("kecamatan", "semua")
    tahun = request.args.get("tahun", "").strip()
    bulan = request.args.get("bulan", "").strip()
    q = request.args.get("q", "").strip()

    conditions = []
    params = []
    if jenis != "semua":
        conditions.append("jenis = ?")
        params.append(jenis)
    if kecamatan != "semua":
        conditions.append("kecamatan = ?")
        params.append(kecamatan)
    if status == "menunggu":
        conditions.append("ditindaklanjuti IS NULL")
    elif status == "ditindaklanjuti":
        conditions.append("ditindaklanjuti = 1 AND instalasi_id IS NULL")
    elif status == "selisih":
        conditions.append("ditindaklanjuti = 0")
    elif status == "selesai":
        conditions.append("instalasi_id IS NOT NULL")
    if tahun:
        conditions.append("strftime('%Y', tanggal_permohonan) = ?")
        params.append(tahun)
    if bulan:
        conditions.append("strftime('%m', tanggal_permohonan) = ?")
        params.append(bulan.zfill(2))
    if q:
        conditions.append("(nama_pelanggan LIKE ? OR lokasi LIKE ? OR kecamatan LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]

    where_clause = " AND ".join(conditions) if conditions else "1=1"
    rows = db.execute(
        f"""SELECT nama_pelanggan, lokasi, kelurahan, kecamatan, jenis,
                          tanggal_permohonan, tanggal_survey, ditindaklanjuti,
                          jenis_pipa, tanggal_dikirim_hublang, tanggal_kembali_hublang,
                          petugas_survey, keterangan
                   FROM permohonan
                   WHERE {where_clause}
                   ORDER BY tanggal_permohonan DESC, id DESC""",
        params,
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Permohonan"
    ws.append([
        "Nama Pelanggan", "Lokasi", "Kelurahan", "Kecamatan", "Jenis",
        "Tanggal Permohonan", "Tanggal Survey", "Ditindaklanjuti", "Jenis Pipa",
        "Tanggal Dikirim Hublang", "Tanggal Kembali Hublang", "Petugas Survey", "Keterangan"
    ])
    for r in rows:
        ws.append([
            r["nama_pelanggan"], r["lokasi"], r["kelurahan"], r["kecamatan"], r["jenis"],
            r["tanggal_permohonan"] or "", r["tanggal_survey"] or "",
            permohonan_ditindaklanjuti_label(r["ditindaklanjuti"]), r["jenis_pipa"] or "",
            r["tanggal_dikirim_hublang"] or "", r["tanggal_kembali_hublang"] or "",
            r["petugas_survey"] or "", r["keterangan"] or ""
        ])

    for idx, width in enumerate([25, 30, 18, 18, 10, 16, 16, 14, 18, 18, 18, 18, 30], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"data_permohonan_{tahun or 'semua'}_{bulan or 'semua'}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/permohonan/kecamatan")
def permohonan_kecamatan():
    db = get_db()
    rows = db.execute(
        """SELECT kecamatan,
                  COUNT(*) total,
                  SUM(CASE WHEN jenis='SIB' THEN 1 ELSE 0 END) sib,
                  SUM(CASE WHEN jenis='BK' THEN 1 ELSE 0 END) bk,
                  SUM(CASE WHEN ditindaklanjuti IS NULL THEN 1 ELSE 0 END) menunggu,
                  SUM(CASE WHEN ditindaklanjuti = 0 THEN 1 ELSE 0 END) selisih,
                  SUM(CASE WHEN instalasi_id IS NOT NULL THEN 1 ELSE 0 END) selesai
           FROM permohonan GROUP BY kecamatan ORDER BY kecamatan"""
    ).fetchall()
    return render_template("permohonan_kecamatan.html", rows=rows)


@app.route("/permohonan/periode")
def permohonan_periode():
    db = get_db()
    tahun_list = [
        r["t"] for r in db.execute(
            """SELECT DISTINCT strftime('%Y', tanggal_permohonan) t FROM permohonan
               WHERE tanggal_permohonan IS NOT NULL AND strftime('%Y', tanggal_permohonan) IS NOT NULL
               ORDER BY t DESC"""
        ).fetchall()
    ]
    tahun_terpilih = request.args.get("tahun", "") or (tahun_list[0] if tahun_list else datetime.now().strftime("%Y"))
    if tahun_list and tahun_terpilih not in tahun_list:
        tahun_terpilih = tahun_list[0]

    per_bulan = db.execute(
        """SELECT strftime('%m', tanggal_permohonan) bulan,
                  COUNT(*) jumlah,
                  SUM(CASE WHEN jenis='SIB' THEN 1 ELSE 0 END) sib,
                  SUM(CASE WHEN jenis='BK' THEN 1 ELSE 0 END) bk
           FROM permohonan
           WHERE strftime('%Y', tanggal_permohonan) = ?
           GROUP BY bulan""",
        (tahun_terpilih,),
    ).fetchall()
    bulan_map = {r["bulan"]: dict(r) for r in per_bulan}
    ringkasan = [
        {
            "nomor": f"{i + 1:02d}", "nama": nama,
            "jumlah": bulan_map.get(f"{i + 1:02d}", {}).get("jumlah", 0),
            "sib": bulan_map.get(f"{i + 1:02d}", {}).get("sib", 0),
            "bk": bulan_map.get(f"{i + 1:02d}", {}).get("bk", 0),
        }
        for i, nama in enumerate(NAMA_BULAN)
    ]
    max_jumlah = max([r["jumlah"] for r in ringkasan], default=1) or 1

    return render_template(
        "permohonan_periode.html", tahun_list=tahun_list, tahun_terpilih=tahun_terpilih,
        ringkasan=ringkasan, max_jumlah=max_jumlah,
    )


@app.route("/permohonan/cari")
def permohonan_cari():
    db = get_db()
    q = request.args.get("q", "").strip()
    scope = request.args.get("scope", "semua")
    rows = []

    if q:
        query = "SELECT * FROM permohonan WHERE 1=1"
        params = []
        if scope == "nama":
            query += " AND nama_pelanggan LIKE ?"
            params.append(f"%{q}%")
        elif scope == "kecamatan":
            query += " AND kecamatan LIKE ?"
            params.append(f"%{q}%")
        else:
            query += " AND (nama_pelanggan LIKE ? OR kecamatan LIKE ? OR lokasi LIKE ?)"
            params += [f"%{q}%", f"%{q}%", f"%{q}%"]
        query += " ORDER BY tanggal_permohonan DESC"
        rows = db.execute(query, params).fetchall()

    return render_template("permohonan_cari.html", q=q, scope=scope, rows=rows)


def hitung_evaluasi_kinerja(db, tahun):
    """Rekap 12 bulan: berapa permohonan masuk, ditindaklanjuti, dan selisih,
    dipecah SIB/BK -- persis struktur "Evaluasi Kinerja Bagian Perencana &
    Supervisi" yang selama ini dibikin manual dari Excel. Dikelompokkan
    berdasarkan bulan tanggal_permohonan masuk (bukan bulan survey)."""
    baris = []
    total = {"masuk_sib": 0, "masuk_bk": 0, "lanjut_sib": 0, "lanjut_bk": 0, "selisih_sib": 0, "selisih_bk": 0}

    for i, nama_bulan in enumerate(NAMA_BULAN, start=1):
        bulan_str = f"{i:02d}"

        def hitung(jenis, kondisi_tambahan=""):
            q = f"""SELECT COUNT(*) c FROM permohonan
                    WHERE jenis=? AND strftime('%Y', tanggal_permohonan)=?
                          AND strftime('%m', tanggal_permohonan)=? {kondisi_tambahan}"""
            return db.execute(q, (jenis, tahun, bulan_str)).fetchone()["c"]

        masuk_sib = hitung("SIB")
        masuk_bk = hitung("BK")
        lanjut_sib = hitung("SIB", "AND ditindaklanjuti=1")
        lanjut_bk = hitung("BK", "AND ditindaklanjuti=1")
        selisih_sib = hitung("SIB", "AND ditindaklanjuti=0")
        selisih_bk = hitung("BK", "AND ditindaklanjuti=0")

        baris.append({
            "bulan": nama_bulan,
            "masuk_sib": masuk_sib, "masuk_bk": masuk_bk,
            "lanjut_sib": lanjut_sib, "lanjut_bk": lanjut_bk,
            "selisih_sib": selisih_sib, "selisih_bk": selisih_bk,
            "total_masuk": masuk_sib + masuk_bk,
            "total_lanjut": lanjut_sib + lanjut_bk,
            "total_selisih": selisih_sib + selisih_bk,
        })
        for k in total:
            total[k] += locals()[k]

    total["total_masuk"] = total["masuk_sib"] + total["masuk_bk"]
    total["total_lanjut"] = total["lanjut_sib"] + total["lanjut_bk"]
    total["total_selisih"] = total["selisih_sib"] + total["selisih_bk"]
    return baris, total


@app.route("/permohonan/laporan")
def permohonan_laporan():
    return redirect(url_for("permohonan_laporan_teknis"))

def ambil_rincian_bulanan(db, tahun, bulan, jenis):
    return db.execute(
        """SELECT nama_pelanggan, lokasi, kelurahan, kecamatan, tanggal_permohonan,
                  tanggal_survey, tanggal_kembali_hublang, petugas_survey, keterangan
           FROM permohonan
           WHERE jenis=? AND strftime('%Y', tanggal_permohonan)=? AND strftime('%m', tanggal_permohonan)=?
           ORDER BY tanggal_permohonan, id""",
        (jenis, tahun, bulan),
    ).fetchall()


def hitung_kinerja_teknik(db, tahun, bulan):
    # 1. Ambil SEMUA keterangan unik dari data bulan terkait
    rows_ket = db.execute(
        """SELECT DISTINCT keterangan FROM permohonan
           WHERE strftime('%Y', tanggal_permohonan)=? AND strftime('%m', tanggal_permohonan)=?
                 AND keterangan IS NOT NULL AND keterangan != ''
           ORDER BY keterangan""",
        (tahun, bulan),
    ).fetchall()
    
    daftar_keterangan = [r["keterangan"] for r in rows_ket]
    
    # Jika pada bulan tersebut belum ada data keterangan, tampilkan daftar standar
    if not daftar_keterangan:
        daftar_keterangan = [
            "Butuh Pipa Distribusi",
            "Meter Air Sudah Terpasang",
            "Bekas Pemutusan / Buka Kembali",
            "Rumah Tidak Ditemukan"
        ]

    hasil = {"daftar_keterangan": daftar_keterangan}

    for jenis in ("SIB", "BK"):
        def hitung(kondisi_tambahan="", params_tambahan=()):
            q = f"""SELECT COUNT(*) c FROM permohonan
                    WHERE jenis=? AND strftime('%Y', tanggal_permohonan)=?
                          AND strftime('%m', tanggal_permohonan)=? {kondisi_tambahan}"""
            return db.execute(q, (jenis, tahun, bulan) + params_tambahan).fetchone()["c"]

        dari_hublang = hitung()
        disurvei = hitung("AND tanggal_survey IS NOT NULL")
        pipa_pendek = hitung("AND jenis_pipa='P.Dinas'")
        pipa_panjang = hitung("AND jenis_pipa='P.Distribusi'")
        dikirim = hitung("AND tanggal_dikirim_hublang IS NOT NULL")

        # Hitung dinamis sesuai keterangan apapun yang ada di database
        detail_keterangan = {}
        total_ket = 0
        for ket in daftar_keterangan:
            c = hitung("AND keterangan=?", (ket,))
            detail_keterangan[ket] = c
            total_ket += c

        hasil[jenis] = {
            "dari_hublang": dari_hublang,
            "ke_perencana": dari_hublang,
            "disurvei": disurvei,
            "pipa_pendek": pipa_pendek,
            "pipa_panjang": pipa_panjang,
            "dikirim": dikirim,
            "detail_keterangan": detail_keterangan,
            "jumlah_ket": total_ket,
        }

    return hasil


@app.route("/permohonan/laporan/teknis")
def permohonan_laporan_teknis():
    db = get_db()
    tahun_list = [
        r["t"] for r in db.execute(
            """SELECT DISTINCT strftime('%Y', tanggal_permohonan) t FROM permohonan
               WHERE tanggal_permohonan IS NOT NULL AND strftime('%Y', tanggal_permohonan) IS NOT NULL
               ORDER BY t DESC"""
        ).fetchall()
    ]
    tahun = request.args.get("tahun", "") or (tahun_list[0] if tahun_list else datetime.now().strftime("%Y"))
    bulan = request.args.get("bulan", "") or datetime.now().strftime("%m")
    nama_bulan = NAMA_BULAN[int(bulan) - 1]

    hasil = hitung_kinerja_teknik(db, tahun, bulan)
    return render_template(
        "permohonan_laporan_teknis.html", tahun_list=tahun_list, tahun=tahun, bulan=bulan,
        nama_bulan=nama_bulan, nama_bulan_list=list(enumerate(NAMA_BULAN, start=1)), 
        hasil=hasil, daftar_keterangan=hasil["daftar_keterangan"]
    )


@app.route("/permohonan/laporan/teknis/unduh")
def permohonan_laporan_teknis_unduh():
    db = get_db()
    tahun = request.args.get("tahun", "")
    bulan = request.args.get("bulan", "")
    nama_bulan = NAMA_BULAN[int(bulan) - 1] if bulan else ""
    
    hasil = hitung_kinerja_teknik(db, tahun, bulan)
    daftar_keterangan = hasil["daftar_keterangan"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Kinerja Supervisi Teknik"

    # Styling Font, Border, Alignment & Fill
    font_title = Font(name="Calibri", bold=True, size=11)
    font_header = Font(name="Calibri", bold=True, size=10)
    font_body = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", bold=True, size=10)
    font_underline = Font(name="Calibri", bold=True, size=10, underline="single")
    
    fill_header = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_side = Side(style="thin", color="000000")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Hitung Jumlah Kolom
    total_cols = 7 + len(daftar_keterangan) + 1
    last_col_letter = get_column_letter(total_cols)

    # 1. Judul Laporan (Baris 1 & 2)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "LAPORAN KINERJA SUPERVISI TEKNIK/GAMBAR/PERHITUNGAN TEKNIS PERMOHONAN SAMBUNGAN INSTALASI BARU DAN BUKA KEMBALI"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"BULAN: {nama_bulan.upper()} {tahun}"
    ws["A2"].font = font_bold
    ws["A2"].alignment = align_center

    # 2. Header Tabel (Baris 4, 5, 6)
    ws.merge_cells("A4:A6")
    ws["A4"] = "Jenis Permohonan"

    ws.merge_cells("B4:C4")
    ws["B4"] = "Permohonan Masuk"

    ws.merge_cells("D4:G4")
    ws["D4"] = "Realisasi Gambar/Perhitungan"

    ket_start_letter = get_column_letter(8)
    ket_end_letter = get_column_letter(total_cols)
    ws.merge_cells(f"{ket_start_letter}4:{ket_end_letter}4")
    ws[f"{ket_start_letter}4"] = "Keterangan (Dikirim ke Hublang)"

    ws.merge_cells("B5:B6")
    ws["B5"] = "Dari Hublang"

    ws.merge_cells("C5:C6")
    ws["C5"] = "Ke Perencana"

    ws.merge_cells("D5:D6")
    ws["D5"] = "Disurvei"

    ws.merge_cells("E5:F5")
    ws["E5"] = "Panjang Pipa"
    ws["E6"] = "≤10 Meter"
    ws["F6"] = ">10m (P.Distribusi)"

    ws.merge_cells("G5:G6")
    ws["G5"] = "Di Kirim ke Hublang"

    col_idx = 8
    for ket in daftar_keterangan:
        col_letter = get_column_letter(col_idx)
        ws.merge_cells(f"{col_letter}5:{col_letter}6")
        ws[f"{col_letter}5"] = ket
        col_idx += 1

    jml_col_letter = get_column_letter(total_cols)
    ws.merge_cells(f"{jml_col_letter}5:{jml_col_letter}6")
    ws[f"{jml_col_letter}5"] = "Jumlah"

    for r in range(4, 7):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all

    # 3. Baris Data (Baris 7 & 8)
    label_jenis = {"SIB": "Sambungan Instalasi Baru (SIB)", "BK": "Buka Kembali (BK)"}
    r = 7
    total_per_col = {c: 0 for c in range(2, total_cols + 1)}

    for jenis in ("SIB", "BK"):
        h = hasil[jenis]
        ws.cell(row=r, column=1, value=label_jenis[jenis]).alignment = align_left
        
        base_vals = [
            h["dari_hublang"], h["ke_perencana"], h["disurvei"],
            h["pipa_pendek"], h["pipa_panjang"], h["dikirim"]
        ]
        
        for idx, val in enumerate(base_vals, start=2):
            val_display = val if val > 0 else "-"
            ws.cell(row=r, column=idx, value=val_display)
            total_per_col[idx] += val

        c_idx = 8
        for ket in daftar_keterangan:
            val = h["detail_keterangan"].get(ket, 0)
            val_display = val if val > 0 else "-"
            ws.cell(row=r, column=c_idx, value=val_display)
            total_per_col[c_idx] += val
            c_idx += 1

        ws.cell(row=r, column=c_idx, value=h["jumlah_ket"] if h["jumlah_ket"] > 0 else "-")
        total_per_col[c_idx] += h["jumlah_ket"]

        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_body
            cell.border = border_all
            if c > 1:
                cell.alignment = align_center
        r += 1

    # 4. Baris Total JUMLAH (Baris 9)
    ws.cell(row=r, column=1, value="JUMLAH").alignment = align_center
    ws.cell(row=r, column=1).font = font_bold

    for c in range(2, total_cols + 1):
        cell = ws.cell(row=r, column=c, value=total_per_col[c])
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_all
    ws.cell(row=r, column=1).border = border_all

    # =========================================================
    # 5. FOOTER TANDA TANGAN DI EXCEL (Sesuai Gambar)
    # =========================================================
    r_ttd = r + 3 # Jarak 3 baris di bawah tabel
    
    # --- SISI KIRI ---
    ws.cell(row=r_ttd, column=2, value="Diperiksa Oleh :").font = font_body
    ws.cell(row=r_ttd + 1, column=2, value="Kabag Perencana & Supervisi").font = font_bold
    
    # Nama Kiri (Bergaris Bawah / Underline)
    c_nama_kiri = ws.cell(row=r_ttd + 5, column=2, value="JIMMI M. SIMATUPANG. ST")
    c_nama_kiri.font = font_underline

    # --- SISI KANAN ---
    col_kanan_start = max(6, total_cols - 3)
    col_kanan_end = total_cols
    
    tgl_hari_ini = datetime.now().strftime("%d %B %Y")
    
    # Kota & Tanggal
    ws.merge_cells(start_row=r_ttd - 1, start_column=col_kanan_start, end_row=r_ttd - 1, end_column=col_kanan_end)
    c_tgl = ws.cell(row=r_ttd - 1, column=col_kanan_start, value=f"Pematangsiantar, {tgl_hari_ini}")
    c_tgl.alignment = align_center
    c_tgl.font = font_body

    # Keterangan & Jabatan Kanan
    ws.merge_cells(start_row=r_ttd, start_column=col_kanan_start, end_row=r_ttd, end_column=col_kanan_end)
    c_ket_k = ws.cell(row=r_ttd, column=col_kanan_start, value="Diperbuat Oleh :")
    c_ket_k.alignment = align_center
    c_ket_k.font = font_body

    ws.merge_cells(start_row=r_ttd + 1, start_column=col_kanan_start, end_row=r_ttd + 1, end_column=col_kanan_end)
    c_jab_k = ws.cell(row=r_ttd + 1, column=col_kanan_start, value="Kasubbag Perencana Jaringan")
    c_jab_k.alignment = align_center
    c_jab_k.font = font_bold

    # Nama Kanan (Bergaris Bawah / Underline)
    ws.merge_cells(start_row=r_ttd + 5, start_column=col_kanan_start, end_row=r_ttd + 5, end_column=col_kanan_end)
    c_nama_kanan = ws.cell(row=r_ttd + 5, column=col_kanan_start, value="SUHAERI")
    c_nama_kanan.alignment = align_center
    c_nama_kanan.font = font_underline

    # 6. Atur Lebar Kolom Otomatis
    ws.column_dimensions["A"].width = 32
    for c in range(2, total_cols + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"laporan_kinerja_teknik_{tahun}-{bulan}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/permohonan/laporan/rincian")
def permohonan_laporan_rincian():
    db = get_db()
    tahun_list = [
        r["t"] for r in db.execute(
            """SELECT DISTINCT strftime('%Y', tanggal_permohonan) t FROM permohonan
               WHERE tanggal_permohonan IS NOT NULL AND strftime('%Y', tanggal_permohonan) IS NOT NULL
               ORDER BY t DESC"""
        ).fetchall()
    ]
    tahun = request.args.get("tahun", "") or (tahun_list[0] if tahun_list else datetime.now().strftime("%Y"))
    bulan = request.args.get("bulan", "") or datetime.now().strftime("%m")
    nama_bulan = NAMA_BULAN[int(bulan) - 1]

    sib_rows = ambil_rincian_bulanan(db, tahun, bulan, "SIB")
    bk_rows = ambil_rincian_bulanan(db, tahun, bulan, "BK")

    return render_template(
        "permohonan_laporan_rincian.html", tahun_list=tahun_list, tahun=tahun, bulan=bulan,
        nama_bulan=nama_bulan, nama_bulan_list=list(enumerate(NAMA_BULAN, start=1)),
        sib_rows=sib_rows, bk_rows=bk_rows,
    )


@app.route("/permohonan/laporan/rincian/unduh")
def permohonan_laporan_rincian_unduh():
    db = get_db()
    tahun = request.args.get("tahun", "")
    bulan = request.args.get("bulan", "")
    nama_bulan = NAMA_BULAN[int(bulan) - 1] if bulan else ""

    headers = ["No", "Nama Pelanggan", "Alamat", "Kelurahan", "Kecamatan",
               "Tgl Ke Perencana", "Tgl Survey", "Kembali Ke Hublang", "Petugas Survey", "Keterangan"]

    def tulis_sheet(ws, judul, rows):
        ws["A1"] = judul
        ws["A1"].font = Font(bold=True, size=12)
        for col, h in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=h).font = Font(bold=True)
        r = 4
        for i, row in enumerate(rows, start=1):
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=row["nama_pelanggan"])
            ws.cell(row=r, column=3, value=row["lokasi"])
            ws.cell(row=r, column=4, value=row["kelurahan"])
            ws.cell(row=r, column=5, value=row["kecamatan"])
            ws.cell(row=r, column=6, value=row["tanggal_permohonan"])
            ws.cell(row=r, column=7, value=row["tanggal_survey"] or "")
            ws.cell(row=r, column=8, value=row["tanggal_kembali_hublang"] or "")
            ws.cell(row=r, column=9, value=row["petugas_survey"] or "")
            ws.cell(row=r, column=10, value=row["keterangan"] or "")
            r += 1
        widths = [5, 22, 26, 14, 14, 14, 12, 14, 14, 20]
        for col, w in zip("ABCDEFGHIJ", widths):
            ws.column_dimensions[col].width = w

    wb = Workbook()
    ws_sib = wb.active
    ws_sib.title = "Permohonan Masuk (SIB)"
    tulis_sheet(ws_sib, f"LAPORAN PERMOHONAN MASUK BULAN {nama_bulan.upper()} {tahun} DARI HUBUNGAN LANGGANAN",
                ambil_rincian_bulanan(db, tahun, bulan, "SIB"))

    ws_bk = wb.create_sheet("Buka Kembali (BK)")
    tulis_sheet(ws_bk, f"LAPORAN PERMOHONAN BUKA KEMBALI BULAN {nama_bulan.upper()} {tahun} DARI HUBUNGAN LANGGANAN",
                ambil_rincian_bulanan(db, tahun, bulan, "BK"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"laporan_permohonan_rincian_{tahun}-{bulan}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/permohonan/laporan/unduh")
def permohonan_laporan_unduh():
    db = get_db()
    tahun = request.args.get("tahun", "")
    baris, total = hitung_evaluasi_kinerja(db, tahun)

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluasi Kinerja"
    bold = Font(bold=True)

    ws["A1"] = "EVALUASI KINERJA BAGIAN PERENCANA & SUPERVISI"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"REKAP SURVEY PASANGAN INSTALASI BARU DAN BUKA KEMBALI - TAHUN {tahun}"

    headers1 = ["No", "Bulan", "Permohonan dari Hublang", "", "Ditindaklanjuti", "", "Selisih", "", "Total"]
    headers2 = ["", "", "SIB", "BK", "SIB", "BK", "SIB", "BK", ""]
    for col, h in enumerate(headers1, start=1):
        ws.cell(row=4, column=col, value=h).font = bold
    for col, h in enumerate(headers2, start=1):
        ws.cell(row=5, column=col, value=h).font = bold

    r = 6
    for i, b in enumerate(baris, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=b["bulan"])
        ws.cell(row=r, column=3, value=b["masuk_sib"])
        ws.cell(row=r, column=4, value=b["masuk_bk"])
        ws.cell(row=r, column=5, value=b["lanjut_sib"])
        ws.cell(row=r, column=6, value=b["lanjut_bk"])
        ws.cell(row=r, column=7, value=b["selisih_sib"])
        ws.cell(row=r, column=8, value=b["selisih_bk"])
        ws.cell(row=r, column=9, value=b["total_masuk"])
        r += 1

    ws.cell(row=r, column=2, value="JUMLAH").font = bold
    for col, key in [(3, "masuk_sib"), (4, "masuk_bk"), (5, "lanjut_sib"), (6, "lanjut_bk"),
                     (7, "selisih_sib"), (8, "selisih_bk"), (9, "total_masuk")]:
        ws.cell(row=r, column=col, value=total[key]).font = bold

    for col, w in [("A", 5), ("B", 12), ("C", 8), ("D", 8), ("E", 8), ("F", 8), ("G", 8), ("H", 8), ("I", 10)]:
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"evaluasi_kinerja_permohonan_{tahun}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/permohonan/tambah", methods=["GET", "POST"])
def permohonan_tambah():
    db = get_db()
    if request.method == "POST":
        errors = validate_permohonan_data(request.form)
        if errors:
            flash("; ".join(errors), "error")
            return render_template("permohonan_form.html", permohonan=None, form_data=request.form,
                                    keterangan_pilihan=KETERANGAN_PERMOHONAN)

        db.execute(
            """INSERT INTO permohonan
               (nama_pelanggan, lokasi, kelurahan, kecamatan, jenis, tanggal_permohonan, keterangan)
               VALUES (?,?,?,?,?,?,?)""",
            (
                request.form["nama_pelanggan"],
                request.form.get("lokasi", ""),
                request.form.get("kelurahan", ""),
                request.form["kecamatan"],
                request.form["jenis"],
                request.form["tanggal_permohonan"],
                request.form.get("keterangan", ""),
            ),
        )
        db.commit()
        flash(f"Permohonan {request.form['nama_pelanggan']} tersimpan.")
        return redirect(url_for("permohonan_list"))

    return render_template("permohonan_form.html", permohonan=None, form_data=None,
                            keterangan_pilihan=KETERANGAN_PERMOHONAN)


@app.route("/permohonan/<int:pmid>/edit", methods=["GET", "POST"])
def permohonan_edit(pmid):
    db = get_db()
    permohonan = db.execute("SELECT * FROM permohonan WHERE id=?", (pmid,)).fetchone()
    if not permohonan:
        flash("Permohonan tidak ditemukan.", "error")
        return redirect(url_for("permohonan_list"))

    if request.method == "POST":
        errors = validate_permohonan_data(request.form)
        if errors:
            flash("; ".join(errors), "error")
            return render_template("permohonan_form.html", permohonan=permohonan, form_data=request.form,
                                    keterangan_pilihan=KETERANGAN_PERMOHONAN)

        ditindaklanjuti_raw = request.form.get("ditindaklanjuti", "")
        ditindaklanjuti = {"ya": 1, "tidak": 0}.get(ditindaklanjuti_raw)  # None kalau "belum"

        db.execute(
            """UPDATE permohonan SET
                 nama_pelanggan=?, lokasi=?, kelurahan=?, kecamatan=?, jenis=?, tanggal_permohonan=?,
                 tanggal_survey=?, petugas_survey=?, ditindaklanjuti=?, jenis_pipa=?,
                 tanggal_dikirim_hublang=?, tanggal_kembali_hublang=?, keterangan=?
               WHERE id=?""",
            (
                request.form["nama_pelanggan"], request.form.get("lokasi", ""),
                request.form.get("kelurahan", ""), request.form["kecamatan"], request.form["jenis"],
                request.form["tanggal_permohonan"], request.form.get("tanggal_survey", "") or None,
                request.form.get("petugas_survey") or None, ditindaklanjuti,
                request.form.get("jenis_pipa") or None, request.form.get("tanggal_dikirim_hublang", "") or None,
                request.form.get("tanggal_kembali_hublang", "") or None,
                request.form.get("keterangan", ""), pmid,
            ),
        )
        db.commit()
        flash("Permohonan diperbarui.")
        return redirect(url_for("permohonan_list"))

    return render_template("permohonan_form.html", permohonan=permohonan, form_data=None,
                            keterangan_pilihan=KETERANGAN_PERMOHONAN)


@app.route("/permohonan/<int:pmid>/jadikan-instalasi", methods=["GET", "POST"])
def permohonan_jadikan_instalasi(pmid):
    db = get_db()
    permohonan = db.execute("SELECT * FROM permohonan WHERE id=?", (pmid,)).fetchone()
    if not permohonan:
        flash("Permohonan tidak ditemukan.", "error")
        return redirect(url_for("permohonan_list"))

    if permohonan["instalasi_id"] is not None:
        flash("Permohonan ini sudah pernah dijadikan instalasi sebelumnya.", "error")
        return redirect(url_for("permohonan_list"))

    if request.method == "POST":
        nomor_instalasi = request.form["nomor_instalasi"].strip()
        dup = db.execute("SELECT 1 FROM instalasi WHERE nomor_instalasi=?", (nomor_instalasi,)).fetchone()
        if dup:
            flash(f"Nomor instalasi \"{nomor_instalasi}\" sudah dipakai instalasi lain.", "error")
            return render_template("permohonan_jadikan_instalasi.html", permohonan=permohonan, form_data=request.form)

        if not is_valid_iso_date(request.form.get("tanggal_pasang", "")):
            flash("Tanggal pasang tidak valid. Pakai date picker, jangan ketik manual.", "error")
            return render_template("permohonan_jadikan_instalasi.html", permohonan=permohonan, form_data=request.form)

        pelanggan_id = cari_atau_buat_pelanggan(
            db, permohonan["nama_pelanggan"], permohonan["lokasi"],
            permohonan["kelurahan"], permohonan["kecamatan"],
        )

        cur = db.execute(
            """INSERT INTO instalasi
               (pelanggan_id, nomor_instalasi, tanggal_pasang, diameter_pipa,
                tekanan_air, status, petugas, keterangan)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                pelanggan_id, nomor_instalasi, request.form["tanggal_pasang"],
                request.form.get("diameter_pipa", ""), request.form.get("tekanan_air") or None,
                permohonan["jenis"], request.form.get("petugas") or None,
                f"Dari permohonan #{pmid}",
            ),
        )
        instalasi_id = cur.lastrowid

        keterangan_baru = permohonan["keterangan"] or "Meter Air Sudah Terpasang"
        db.execute(
            "UPDATE permohonan SET instalasi_id=?, keterangan=? WHERE id=?",
            (instalasi_id, keterangan_baru, pmid),
        )
        db.commit()
        flash(f"Instalasi {nomor_instalasi} dibuat dari permohonan {permohonan['nama_pelanggan']}.")
        return redirect(url_for("permohonan_list"))

    return render_template("permohonan_jadikan_instalasi.html", permohonan=permohonan, form_data=None)


@app.route("/permohonan/<int:pmid>/hapus", methods=["POST"])
def permohonan_hapus(pmid):
    db = get_db()
    permohonan = db.execute("SELECT nama_pelanggan, instalasi_id FROM permohonan WHERE id=?", (pmid,)).fetchone()
    if not permohonan:
        flash("Permohonan tidak ditemukan.", "error")
        return redirect(url_for("permohonan_list"))

    db.execute("DELETE FROM permohonan WHERE id=?", (pmid,))
    db.commit()

    if permohonan["instalasi_id"]:
        flash(f"Permohonan {permohonan['nama_pelanggan']} dihapus. Instalasi yang sudah terlanjur dibuat dari permohonan ini TIDAK ikut terhapus.")
    else:
        flash(f"Permohonan {permohonan['nama_pelanggan']} dihapus.")
    return redirect(url_for("permohonan_list"))


if __name__ == "__main__":
    ensure_data_paths()

    # Otomatis buka browser saat file .exe dijalankan
    import webbrowser
    from threading import Timer

    def open_browser():
        webbrowser.open("http://127.0.0.1:5000")

    # Jalankan pembukaan browser dengan jeda 1.5 detik agar Flask siap mendengarkan port
    Timer(1.5, open_browser).start()

    # Matikan debug mode (debug=False) saat di-package ke exe agar tidak memicu reloading ganda
    # Mode ditentukan dari luar (run.bat), bukan hardcode di sini -- biar
    # app.py ini satu-satunya sumber kebenaran buat semua varian (local,
    # shared, Win7, Win10). Yang beda antar folder cukup run.bat-nya doang.
    sidap_mode = os.environ.get("SIDAP_MODE", "local")
    host = "0.0.0.0" if sidap_mode == "shared" else "127.0.0.1"
    print(f"SIDAP jalan dalam mode: {sidap_mode} (host={host})")

    app.run(debug=False, host=host, port=5000, threaded=True)